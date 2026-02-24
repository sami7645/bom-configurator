"""
Convert relevant sheets from the master Excel file to CSV files
used by the BOM Configurator's import_csv_data command.

Usage:
    python convert_excel_to_csv.py
    python convert_excel_to_csv.py "path/to/excel_file.xlsx"
    python convert_excel_to_csv.py "path/to/excel_file.xlsx" --output csv_files
    python convert_excel_to_csv.py "path/to/excel_file.xlsx" --output csv_files --sheet DFM

After conversion, re-import into DB:
    python manage.py import_csv_data --force
"""
import argparse
import csv
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# Map: Excel sheet name → CSV filename expected by import_csv_data.py
SHEET_TO_CSV = {
    "Schacht": "Schacht.csv",
    "Schachtgrenze": "Schachtgrenze.csv",
    "Schachtkompatibilität": "Schachtkompatibilitaet.csv",
    "HVB": "HVB.csv",
    "Sondenabstaende": "Sondenabstaende.csv",
    "Stumpfschweiß-Endkappen": "Stumpfschweiss-Endkappen.csv",
    "WPA": "WPA.csv",
    "WP-Verschlusskappen": "WP-Verschlusskappen.csv",
    "Kugelhähne": "Kugelhaehne.csv",
    "DFM": "DFM.csv",
    "Sondengröße - Sondenlänge": "Sondengroesse - Sondenlaenge.csv",
    "Sonden-Durchmesser": "Sonden-Durchmesser.csv",
    "Schacht-Sondendurchmesser": "Schacht-Sondendurchmesser.csv",
    "GN X - Articles": "GN X - Articles.csv",
    "Sondenverschlusskappe": "Sondenverschlusskappe.csv",
    "Entlüftung": "Entlueftung.csv",
    "Verrohrung": "Verrohrung.csv",
}


def format_cell(value):
    """Format a single cell value for CSV output."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    return str(value)


def sheet_to_csv(ws, csv_path):
    """Write an openpyxl worksheet to a CSV file (UTF-8 with BOM)."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"  SKIP (empty sheet)")
        return 0

    headers = rows[0]
    # Strip trailing None columns from header
    while headers and headers[-1] is None:
        headers = headers[:-1]
    num_cols = len(headers)

    data_rows = rows[1:]
    written = 0

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([format_cell(h) for h in headers])
        for row in data_rows:
            trimmed = row[:num_cols]
            formatted = [format_cell(c) for c in trimmed]
            # Skip completely empty rows
            if all(v == "" for v in formatted):
                continue
            writer.writerow(formatted)
            written += 1

    return written


def convert(excel_path: str, output_dir: str, sheet_only: str = None):
    """Convert all relevant sheets from *excel_path* into *output_dir*.
    If sheet_only is set (e.g. 'DFM'), only that sheet is converted."""
    print(f"Loading workbook: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    os.makedirs(output_dir, exist_ok=True)
    print(f"Output directory : {output_dir}\n")

    converted = []
    items = list(SHEET_TO_CSV.items())
    if sheet_only:
        items = [(s, c) for s, c in items if s == sheet_only]
        if not items:
            print(f"  ERROR: Sheet '{sheet_only}' not in SHEET_TO_CSV map.")
            wb.close()
            return []

    for sheet_name, csv_name in items:
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet '{sheet_name}' not found in workbook - skipped")
            continue

        ws = wb[sheet_name]
        csv_path = os.path.join(output_dir, csv_name)
        print(f"  {sheet_name:40s} -> {csv_name}")
        count = sheet_to_csv(ws, csv_path)
        print(f"     {count} data rows written")
        converted.append(csv_name)

    wb.close()
    return converted


def copy_manual_csvs(source_dir: str, output_dir: str):
    """Copy CSV files that are NOT in the Excel (manually maintained)."""
    manual_files = [
        "GNXChamberArticle.csv",
        "AdditionalProbeCombinations.csv",
    ]
    copied = []
    for fname in manual_files:
        src = os.path.join(source_dir, fname)
        dst = os.path.join(output_dir, fname)
        if os.path.exists(src):
            import shutil
            shutil.copy2(src, dst)
            print(f"  Copied manual CSV: {fname}")
            copied.append(fname)
        else:
            print(f"  WARNING: Manual CSV not found: {src}")
    return copied


if __name__ == "__main__":
    base_dir = Path(__file__).resolve().parent

    # Default Excel file (newest V3)
    default_excel = None
    for candidate in [
        "main excel/26.02.10 - Stammdaten_Regelwerk_Lookups_Einzelteile_V3 (1).xlsx",
        "26.02.10 - Stammdaten_Regelwerk_Lookups_Einzelteile_V3.xlsx",
    ]:
        if (base_dir / candidate).exists():
            default_excel = str(base_dir / candidate)
            break

    parser = argparse.ArgumentParser(description="Convert Excel sheets to CSV")
    parser.add_argument("excel_file", nargs="?", default=default_excel, help="Path to Excel file")
    parser.add_argument("--output", "-o", default=str(base_dir / "csv_files"), help="Output directory for CSV files")
    parser.add_argument("--sheet", "-s", default=None, help="Convert only this sheet (e.g. DFM)")
    args = parser.parse_args()

    excel_path = args.excel_file
    output_dir = str(Path(args.output).resolve())
    sheet_only = args.sheet

    if not excel_path or not os.path.exists(excel_path):
        print("ERROR: Excel file not found. Pass the path as first argument.")
        print("Usage: python convert_excel_to_csv.py <excel_file> [--output csv_files] [--sheet DFM]")
        sys.exit(1)

    print("=" * 60)
    print("  Excel -> CSV Conversion")
    if sheet_only:
        print(f"  (single sheet: {sheet_only})")
    print("=" * 60)

    converted = convert(excel_path, output_dir, sheet_only=sheet_only)

    print()
    # Manual CSVs: only copy when doing full convert and output is a different dir.
    old_csv_dir = str((base_dir / "csv_files").resolve())
    if not sheet_only and old_csv_dir != output_dir:
        copied = copy_manual_csvs(old_csv_dir, output_dir)
    else:
        copied = []
        print("  Manual CSVs already in output directory - skipping copy")

    print()
    print("=" * 60)
    total = len(converted) + len(copied)
    print(f"  Done! {total} CSV files written to {output_dir}")
    print()
    print("  Next step – re-import into database:")
    print("    python manage.py import_csv_data --force")
    print("=" * 60)
